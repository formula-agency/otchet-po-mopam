from __future__ import annotations

import argparse
import json
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


SHEETS = {
    "Высокий приоритет 19.08": ("2026-08-19", 0, 1),
    "20.08": ("2026-08-20", 0, 1),
    "21.08": ("2026-08-21", 0, 1),
    "24.08": ("2026-08-24", 0, 1),
    "25.08 перетечение изо дня в ден": ("2026-08-25", 1, 0),
}
MOP_NAMES = {
    "тончу ростислав федорович": "Тончу Ростислав",
    "жуков лев сергеевич": "Жуков Лев",
    "попова олеся владимировна": "Попова Олеся",
    "гавриленко елена ивановна": "Гавриленко Елена",
}


def deal_id_from_link(value: Any) -> str:
    match = re.search(r"/details/(\d+)", str(value or ""))
    return match.group(1) if match else ""


def canonical_mop_name(value: Any) -> str:
    raw_name = " ".join(str(value or "").split())
    return MOP_NAMES.get(raw_name.casefold(), raw_name)


def optional_int(value: Any) -> int | None:
    if value is None or value == "":
        return None
    return int(float(value))


def read_snapshot(worksheet: Any, snapshot_date: str, name_col: int, link_col: int) -> dict[str, Any]:
    deals: dict[str, dict[str, Any]] = {}
    detailed = snapshot_date == "2026-08-25"
    for values in worksheet.iter_rows(min_row=2, values_only=True):
        deal_id = deal_id_from_link(values[link_col] if link_col < len(values) else "")
        mop_name = canonical_mop_name(values[name_col] if name_col < len(values) else "")
        if not deal_id or not mop_name:
            continue
        stage_name = str(values[2] or "") if detailed and len(values) > 2 else ""
        days_without_attempt = optional_int(values[3]) if detailed and len(values) > 3 else None
        days_without_call = optional_int(values[4]) if detailed and len(values) > 4 else None
        deals[deal_id] = {
            "dealId": deal_id,
            "dealUrl": f"https://crm.formula-agency.com/crm/deal/details/{deal_id}/",
            "title": "",
            "mopId": "",
            "mopName": mop_name,
            "stageId": "",
            "stageName": stage_name,
            "dateCreate": "",
            "lastCallAttemptDate": "",
            "lastSuccessfulCommunicationDate": "",
            "daysWithoutAttempt": days_without_attempt,
            "daysWithoutCall": days_without_call,
        }
    return {
        "date": snapshot_date,
        "deals": sorted(
            deals.values(),
            key=lambda row: (
                -(row.get("daysWithoutCall") or 0),
                row["mopName"],
                int(row["dealId"]),
            ),
        ),
    }


def main() -> int:
    parser = argparse.ArgumentParser(description="Импортирует ежедневные снимки TempLab из XLSX.")
    parser.add_argument("workbook")
    parser.add_argument(
        "--output",
        default="manual-data/high-priority-history.json",
    )
    args = parser.parse_args()

    workbook = load_workbook(args.workbook, data_only=True, read_only=True)
    snapshots: dict[str, Any] = {}
    previous_date = ""
    for sheet_name, (snapshot_date, name_col, link_col) in SHEETS.items():
        if sheet_name not in workbook.sheetnames:
            continue
        snapshot = read_snapshot(workbook[sheet_name], snapshot_date, name_col, link_col)
        current_ids = {row["dealId"] for row in snapshot["deals"]}
        previous_ids = {
            row["dealId"] for row in snapshots.get(previous_date, {}).get("deals", [])
        }
        snapshot.update({
            "previousDate": previous_date,
            "calledFromPreviousDealIds": sorted(previous_ids - current_ids, key=int),
            "flowedFromPreviousDealIds": sorted(previous_ids & current_ids, key=int),
            "newOverdueDealIds": sorted(current_ids - previous_ids, key=int),
            "calledFromPreviousEvaluated": bool(previous_date),
            "calledFromPreviousSource": "templab-daily-diff",
        })
        snapshots[snapshot_date] = snapshot
        previous_date = snapshot_date

    output = Path(args.output)
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(
        json.dumps({"schemaVersion": 4, "snapshots": snapshots}, ensure_ascii=False, indent=2)
        + "\n",
        encoding="utf-8",
    )
    print(f"Imported {len(snapshots)} TempLab snapshots to {output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
